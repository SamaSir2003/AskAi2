import os
from flask import Flask, request, render_template, jsonify
from dotenv import load_dotenv
from PyPDF2 import PdfReader  # Import for reading PDF files
from io import BytesIO
import mimetypes
from openpyxl import load_workbook
from docx import Document
from langchain_google_genai import ChatGoogleGenerativeAI

load_dotenv()
api_key = os.getenv('api_key')

llm = ChatGoogleGenerativeAI(
    model="gemini-1.5-pro",
    api_key=api_key
)


app = Flask(__name__)

#genai.configure(api_key=api_key)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_file():
    file = request.files['file']
    question = request.form.get('question')

    if not file or not question:
        return render_template('process.html')

    try:
        file_content = ""
        

        if file.mimetype == 'application/pdf':  # If the file is a PDF
            reader = PdfReader(BytesIO(file.read()))
            for page in reader.pages:
                file_content += page.extract_text()
                
        elif "xls" in file.mimetype or "sheet" in file.mimetype or "xl" in file.mimetype or "excel" in file.mimetype:
            # Use pandas to read Excel file
            
            workbook = load_workbook(BytesIO(file.read()), read_only=True)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    file_content += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"

        elif file.mimetype == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':  # For .docx
            document = Document(BytesIO(file.read()))
            for paragraph in document.paragraphs:
                file_content += paragraph.text + "\n"

        else:  # If the file is a text file
            print("Hello excel")
                
            file_content = file.read().decode('utf-8')
    except Exception as e:
        return render_template('index.html', error_message=f"Failed to process the file: {e}")

    # Use Generative Model to generate a response
    #model = genai.GenerativeModel("gemini-1.5-flash")
    try:
        # response = model.generate_content(
        #     f"For the content of a data given below\n{file_content}\nAnswer the question below. If there are multiple records, separate them with a newline\n{question}"
        # )
        response=llm.invoke(f"For the content of a data given below\n{file_content}\nAnswer the question below. If there are multiple records, separate them with a newline\n{question}")
        answer = response.content
    except Exception as e:
        
        return render_template('index.html', error_message=f"Failed to generate a response: {e}")

    return render_template("process.html", question=question, answer=answer)

if __name__ == '__main__':
    app.run(debug=True)
